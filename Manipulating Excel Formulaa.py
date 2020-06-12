# -*- coding: utf-8 -*-
"""
Created on Tue Apr  7 11:44:29 2020


"""

import openpyxl
import os 
from openpyxl.formula import Tokenizer
from openpyxl.formula.translate import Translator


os.chdir("D:\Roisin")

#Read in workbook
wb = openpyxl.load_workbook('RoisinExcel.xlsx')

#Get sheet names
sheetnames = wb.sheetnames

#Select sheet of interest
c_sheet = wb[sheetnames[1]]

#Select column of interest
c_sheet["B"]

#Select first sheet
first_sheet = wb[sheetnames[0]]



#TO do: Figure out how to use dynamic cell references rather than hard coded as below.
for row in range(1,c_sheet.max_row +1)

cellref = "B1"

# Translate formula with autoincrements
 c_sheet.cell(row,3).value = Translator(c_sheet.cell(row,2).value, origin = "B1").translate_formula("C1")
 
 #This will do the same as above but with dynamic cell refernce.
 c_sheet.cell(row,3).value = Translator(c_sheet.cell(row,2).value, origin = cellref).translate_formula("C1")


# Copy formula as is.  
 c_sheet.cell(2,3).value = c_sheet.cell(2,2).value
 

 


